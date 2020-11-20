#!/usr/bin/env python3
# -*- coding:utf-8 -*-
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from config.conf import EXCEL_PATH


class ParseExcel(object):
    def __init__(self, in_excel_name):
        self.excel_name = '%s.xlsx' % in_excel_name
        self.excelPath = os.path.join(EXCEL_PATH, self.excel_name)
        # 加载excel
        self.workbook = load_workbook(self.excelPath)
        # 获取第一个sheet页
        self.sheet = self.workbook.active
        self.font = Font(color=None)
        self.colorDict = {"red": 'FFFF3030', "green": 'FF008B00'}

    # 设置当前要操作的sheet对象，使用index获取相应的sheet
    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)

        return self.sheet

    # 获取当前默认sheet的名字
    def get_default_sheet(self):
        return self.sheet.title

    # 设置当前要操作的sheet对象，使用sheet名称获取相应的sheet
    def set_sheet_by_name(self, sheet_name):
        self.sheet = self.workbook[sheet_name]

        return self.sheet

    # 获取默认sheet中最大行数
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取默认sheet中最大列数
    def get_max_col_no(self):
        return self.sheet.max_column

    # 获取默认sheet中的最小（起始）行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet中的最小（起始）列号
    def get_min_col_no(self):
        return self.sheet.min_column

    # 获取所有默认sheet的所有行对象
    def get_all_rows(self):
        return list(self.sheet.iter_rows())

    # 获取所有默认sheet的所有列对象
    def get_all_cols(self):
        return list(self.sheet.iter_cols())

    # 从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    # 从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    # 从默认sheet中，通过行号和列号获取指定单元格，注意行号和列号要从1开始
    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    # 从默认sheet中，通过行号和列号获取指定单元格中的内容，注意行号和列号要从1开始
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始
    # 调用此方法时，excel不要处于打开状态
    def write_cell_content(self, row_no, col_no, content):
        self.sheet.cell(row=row_no, column=col_no).value = content
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始
    # 调用此方法时，excel不要处于打开状态
    def write_cell_current_time(self, row_no, col_no):
        time1 = time.strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.cell(row=row_no, column=col_no).value = str(time1)
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=col_no).value

    def save_excel_file(self):
        self.workbook.save(self.excelPath)

    @staticmethod
    def get_excel_data(bd, sheet_name):
        bd.set_sheet_by_name(sheet_name)
        max_row = bd.get_max_row_no()
        max_cell = bd.get_max_col_no()
        data = {}
        td = []
        for row in range(2, max_row + 1):
            for cell in range(1, max_cell + 1):
                key = bd.get_cell_content(1, cell)
                value = bd.get_cell_content(row, cell)
                data[key] = value
            td.append(data)
        return td
